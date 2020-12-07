import openpyxl
#import os 
#os.getcwd()
#os.chdir() 

import pandas as pd 
file = '/home/lohith/Desktop/EY Hackathon/EY_GDS_Project/data_to_preprocess/Project Management/Excel-Project-Tracker.xlsx'
# Read and store content 
# of an excel file  
read_file = pd.read_excel (file) 
  
# Write the dataframe object 
# into csv file 
read_file.to_csv ("Excel.csv",  
                  index = None, 
                  header=True) 
    
# read csv file and convert  
# into a dataframe object 
df = pd.DataFrame(pd.read_csv("Excel.csv")) 
  
# show the dataframe 
print(df) 
# data = pd.ExcelFile(file)
# print(data.sheet_names) 

# df = data.parse('Sheet1')
# df.info
# df.head(10)

# ps = openpyxl.load_workbook('produceSales.xlsx')
# sheet = ps['Sheet1']
# sheet.max_row 


# for row in range(2, sheet.max_row + 1):
#  produce = sheet['B' + str(row)].value
#  cost_per_pound = sheet['C' + str(row)].value
#  pounds_sold = sheet['D' + str(row)].value
#  total_sales = sheet['E' + str(row)].value

# TotalInfo={}
# TotalInfo.setdefault(produce,{'Total_cost_per_pound': 0,'Total_pounds_sold': 0, 'Total_sales': 0,'Total_Purchase_Instances': 0})
# TotalInfo[produce]['Total_cost_per_pound'] += float(cost_per_pound)
# TotalInfo[produce]['Total_pounds_sold'] += int(pounds_sold)
# TotalInfo[produce]['Total_sales'] += int(total_sales)
# TotalInfo[produce]['Total_Purchase_Instances'] += 1


# resultFile = open('Total_info.txt', 'w')
# resultFile.write(print.pformat(TotalInfo))
# resultFile.close()
# print('Done.')
# Open ('Total_info.csv', 'w')
